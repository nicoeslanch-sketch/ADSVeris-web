from html import escape
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Balance_General" / "Balance_General.xlsx"
OUT_DIR = ROOT / "assets" / "productos" / "balance"

SHEETS = [
    ("Inicio", "1.svg", "Vista inicial"),
    ("Balance General", "2.svg", "Balance horizontal"),
    ("Dashboard", "3.svg", "Dashboard"),
    ("Activos", "4.svg", "Activos"),
    ("Pasivos", "5.svg", "Pasivos"),
]


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        if abs(value) >= 1000:
            return f"${value:,.0f}".replace(",", ".")
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value)


def relevant_rows(ws, max_rows=13, max_cols=6):
    rows = []
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row or max_rows, 80),
        min_col=1,
        max_col=min(ws.max_column or max_cols, max_cols),
        values_only=True,
    ):
        values = [cell_text(value).strip() for value in row]
        if any(values):
            rows.append(values)
        if len(rows) >= max_rows:
            break
    return rows


def row_svg(values, y, index):
    cells = []
    widths = [260, 260, 260, 180, 140, 180]
    x = 74
    for col, text in enumerate(values[:6]):
        width = widths[col] if col < len(widths) else 160
        fill = "#f7fbff" if index == 0 else ("#e9f6f2" if index % 2 else "#ffffff")
        text_fill = "#0d2235" if index == 0 else "#243447"
        weight = "800" if index == 0 else "600"
        cells.append(
            f'<rect x="{x}" y="{y}" width="{width}" height="44" fill="{fill}" stroke="#d9e2ea" />'
        )
        if text:
            cells.append(
                f'<text x="{x + 14}" y="{y + 27}" fill="{text_fill}" font-size="15" '
                f'font-weight="{weight}">{escape(text[:32])}</text>'
            )
        x += width
    return "\n".join(cells)


def make_svg(sheet_name, label, rows):
    title = rows[0][0] if rows and rows[0] else sheet_name
    subtitle = rows[1][0] if len(rows) > 1 and rows[1] else label
    body_rows = rows[2:] if len(rows) > 2 else rows
    rendered_rows = "\n".join(row_svg(row, 238 + index * 44, index) for index, row in enumerate(body_rows[:9]))

    return f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 1200 760" role="img" aria-label="{escape(sheet_name)}">
  <defs>
    <linearGradient id="bg" x1="0" x2="1" y1="0" y2="1">
      <stop offset="0" stop-color="#0b2538"/>
      <stop offset="0.58" stop-color="#143f52"/>
      <stop offset="1" stop-color="#0f766e"/>
    </linearGradient>
    <filter id="shadow" x="-20%" y="-20%" width="140%" height="140%">
      <feDropShadow dx="0" dy="22" stdDeviation="24" flood-color="#03111f" flood-opacity="0.28"/>
    </filter>
  </defs>
  <rect width="1200" height="760" rx="34" fill="url(#bg)"/>
  <circle cx="1010" cy="140" r="170" fill="#c9a84c" opacity="0.16"/>
  <circle cx="160" cy="665" r="210" fill="#54d89d" opacity="0.14"/>
  <g filter="url(#shadow)">
    <rect x="54" y="54" width="1092" height="652" rx="24" fill="#f8fbfc"/>
    <rect x="54" y="54" width="1092" height="66" rx="24" fill="#10283d"/>
    <rect x="54" y="96" width="1092" height="24" fill="#10283d"/>
    <circle cx="92" cy="87" r="9" fill="#f06a5b"/>
    <circle cx="124" cy="87" r="9" fill="#f7c75f"/>
    <circle cx="156" cy="87" r="9" fill="#54d89d"/>
    <text x="204" y="92" fill="#f8fbfc" font-family="Arial, sans-serif" font-size="22" font-weight="800">Balance_General.xlsx</text>
    <text x="74" y="164" fill="#0f766e" font-family="Arial, sans-serif" font-size="16" font-weight="800" letter-spacing="2">{escape(label.upper())}</text>
    <text x="74" y="204" fill="#0d2235" font-family="Arial, sans-serif" font-size="34" font-weight="900">{escape(title[:58])}</text>
    <text x="74" y="228" fill="#526173" font-family="Arial, sans-serif" font-size="15" font-weight="600">{escape(subtitle[:108])}</text>
    <g font-family="Arial, sans-serif">{rendered_rows}</g>
  </g>
</svg>
'''


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(WORKBOOK, data_only=True, read_only=True)
    for sheet_name, filename, label in SHEETS:
        ws = workbook[sheet_name]
        rows = relevant_rows(ws)
        (OUT_DIR / filename).write_text(make_svg(sheet_name, label, rows), encoding="utf-8")


if __name__ == "__main__":
    main()
